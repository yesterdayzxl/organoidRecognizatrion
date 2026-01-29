from ultralytics import YOLO
import cv2
import os
import numpy as np
import openpyxl
import re
from tqdm import tqdm

# 加载预训练的 YOLOv8 模型
model = YOLO(r'F:\Projects\CellImaging\cellImaging\Project\runs\segment\train72\weights\best.pt')

# Excel 输出路径和文件名
excel_output_dir = 'F:\\Picture05'
excel_name = "results.xlsx"
excel_path = os.path.join(excel_output_dir, excel_name)

# 如果 Excel 文件不存在，创建一个新文件
if not os.path.exists(excel_path):
    wb = openpyxl.Workbook()
    wb.save(excel_path)


# 颜色映射 (hex to BGR)
def hex_to_bgr(hex):
    return tuple(int(hex[i:i + 2], 16) for i in (4, 2, 0))


hexs = (
    "FF3838", "FF9D97", "FF701F", "FFB21D", "CFD231",
    "48F90A", "92CC17", "3DDB86", "1A9334", "00D4BB",
    "2C99A8", "00C2FF", "344593", "6473FF", "0018EC",
    "8438FF", "520085", "CB38FF", "FF95C8", "FF37C7"
)
colors = [hex_to_bgr(h) for h in hexs]


# 非极大值抑制 (NMS)
def nms(boxes, confidences, iou_threshold=0.5):
    if len(boxes) == 0:
        return []

    boxes = np.array(boxes)
    confidences = np.array(confidences)

    areas = (boxes[:, 2] - boxes[:, 0] + 1) * (boxes[:, 3] - boxes[:, 1] + 1)
    order = confidences.argsort()[::-1]
    keep = []

    while order.size > 0:
        i = order[0]
        keep.append(i)

        xx1 = np.maximum(boxes[i, 0], boxes[order[1:], 0])
        yy1 = np.maximum(boxes[i, 1], boxes[order[1:], 1])
        xx2 = np.minimum(boxes[i, 2], boxes[order[1:], 2])
        yy2 = np.minimum(boxes[i, 3], boxes[order[1:], 3])

        w = np.maximum(0, xx2 - xx1 + 1)
        h = np.maximum(0, yy2 - yy1 + 1)
        inter = w * h
        iou = inter / (areas[i] + areas[order[1:]] - inter)

        inds = np.where(iou <= iou_threshold)[0]
        order = order[inds + 1]

    return keep


# 文件名提取分组和天数信息
def extract_group_and_day(filename):
    match = re.match(r'(\d+)D(\d+)', filename)
    if match:
        group = int(match.group(1))
        day = int(match.group(2))
        return group, day
    return None, None


# def extract_group_and_day(filename):
#     match = re.match(r'([A-Za-z]+)(\d+)', filename)  # + 表示一个或多个字母
#     if match:
#         letters = match.group(1)  # 提取字母部分
#         number = int(match.group(2))  # 提取数字部分
#         return letters, number
#     return None, None

# 获取文件的基本名
def get_last_part_of_string(path):
    return os.path.basename(path)


def ensure_3_channels(image):
    """确保图像是3通道BGR格式，适用于各种输入情况"""
    if image is None:
        return None

    # 如果是16位图像，先转换为8位
    if image.dtype == np.uint16:
        image = (image / 256).astype(np.uint8)

    # 检查通道数
    if len(image.shape) == 2:
        # 灰度图，转换为3通道
        image = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
    elif len(image.shape) == 3:
        if image.shape[2] == 1:
            # 单通道图像，转换为3通道
            image = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
        elif image.shape[2] == 4:
            # RGBA图像，转换为BGR
            image = cv2.cvtColor(image, cv2.COLOR_BGRA2BGR)
        # 如果已经是3通道，不需要转换

    return image


# 遍历 chip 编号
for chip_no in range(1, 2):
    source_dir = f'F:\\Picture05\\chip{chip_no}\\'
    gray_source_dir = f'F:\\Picture05\\chip{chip_no}-matched\\'
    output_dir = f'F:\\Picture05\\chip{chip_no}-res\\'

    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 打开 Excel 文件并检查是否存在当前 chip 的工作表
    wb = openpyxl.load_workbook(excel_path)
    sheet_name = f'chip{chip_no}result'
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Group", "Day", "Image Name", "Class Name", "Area", "Confidence", "Gray Mean", "Gray StdDev"])
    else:
        ws = wb[sheet_name]

    detection_results = []

    # 获取所有文件列表
    all_files = [
        os.path.join(root, file)
        for root, _, files in os.walk(gray_source_dir)
        for file in files
        if file.endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff'))
    ]

    # 设置进度条
    with tqdm(total=len(all_files), desc=f"Processing chip {chip_no}", unit="file") as pbar:
        for source in all_files:
            file = get_last_part_of_string(source)
            group, day = extract_group_and_day(file)
            if group is None or day is None:
                pbar.write(f"Skipped file without valid format: {file}")
                pbar.update(1)
                continue

            # 读取图像（使用 IMREAD_UNCHANGED 保留原始格式）
            img_raw = cv2.imread(source, cv2.IMREAD_UNCHANGED)
            if img_raw is None:
                pbar.write(f"Failed to read image: {source}")
                pbar.update(1)
                continue

            # 确保转换为3通道BGR格式
            img_color = ensure_3_channels(img_raw)
            if img_color is None:
                pbar.write(f"Failed to convert image: {source}")
                pbar.update(1)
                continue

            # 预测（使用3通道图像）
            results = model(source=img_color, save=True, retina_masks=True)

            for result in results:
                if not result.boxes or not result.masks:
                    pbar.write(f"No detections in: {source}")
                    continue

                image_path = gray_source_dir + file
                image_gray = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
                # image_gray_bg = image_gray[0:32,0:32]
                boxes = result.boxes.xyxy.cpu().numpy()
                confidences = result.boxes.conf.cpu().numpy()
                names = result.names

                # 筛选置信度大于 0.6 的目标
                high_conf_indices = np.where(confidences > 0.6)[0]

                for idx in high_conf_indices:
                    box = boxes[idx]
                    confidence = confidences[idx]
                    class_id = int(result.boxes.cls[idx])
                    color = colors[class_id % len(colors)]
                    mask = (result.masks.data[idx].cpu().numpy() > 0.95).astype(np.uint8)
                    mask_bg = np.ones((32, 32)).astype(np.uint8)
                    # 计算面积和灰度信息
                    area = np.sum(mask)
                    mean_val, std_val = cv2.meanStdDev(image_gray, mask=mask)
                    # bg_mean_val, bg_std_val = cv2.meanStdDev(image_gray_bg, mask = mask_bg)
                    detection_results.append([
                        group,
                        day,
                        file,
                        names[class_id],
                        round(area, 2),
                        round(confidence, 2),
                        round(mean_val[0][0], 2),
                        round(std_val[0][0], 2),
                        # round(bg_mean_val[0][0], 2),  # 添加背景平均灰度值
                        # round(bg_std_val[0][0], 2)  # 添加背景灰度标准差
                    ])

                    # 绘制轮廓
                    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    for contour in contours:
                        cv2.drawContours(image_gray, [contour], -1, color, 2)

                    # 添加信息标注
                    text = (f'{names[class_id]} Area: {area:.2f} Conf: {confidence:.2f} '
                            f'Mean: {mean_val[0][0]:.2f} Std: {std_val[0][0]:.2f}')
                    cv2.putText(image_gray, text, (int(box[0]), int(box[1]) - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

                annotated_output_path = os.path.join(output_dir, get_last_part_of_string(image_path))
                cv2.imwrite(annotated_output_path, image_gray)

            # 更新进度条
            pbar.update(1)

    # 按分组和天数排序并写入 Excel
    detection_results.sort(key=lambda x: (x[0], x[1]))
    for row in detection_results:
        ws.append(row)

    # 保存当前工作表
    wb.save(excel_path)
    print(f"Results for chip{chip_no} saved in sheet '{sheet_name}'.")

print(f"All results saved in Excel file: {excel_path}")